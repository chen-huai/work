#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import os.path
import time
import easygui as g
from openpyxl import Workbook
from openpyxl import load_workbook
import random
import csv
import tkinter
now_1 = str(time.strftime('%Y'))
def get_sample():
    i = 0
    # path=tkinter.filedialog.askopenfilenames(initialdir='Z:/Inorganic_batch/Formaldehyde/batch/*.doc*',filetypes=[("wordfile", "*.doc*")])
    path = g.fileopenbox(msg=None, title=None, default='Z:/Inorganic_batch/Formaldehyde/batch/*', filetypes=None)
    print(path)
    sequence_mm = []
    for file in path:
        print(file,1)
        w = Dispatch('Word.Application')
        w.Visible = 0
        doc = w.Documents.Open(r'%s'%file, Encoding='utf-8')
        a = doc.Content.Text
        b = a.split('\r')
        doc.Close()




def get_list():  
    lab_number = input('please fill in the lab number:')
    quality=float(input('please fill in the sample quality:'))
    volume=float(input('please fill in the volume of the sample:'))
    print ('begin')
    print ('Wait for a moment')
    name = lab_number.replace('/', '_')
    lines = csv.reader(open('%s'%path,'rt',encoding='utf-8'))
    lists = []
    for l in lines:
        lists.append(l)
    workbook = load_workbook(os.path.join(os.getcwd(),r'./SVHC.xlsx'))
    sheets_1 = workbook.sheetnames
    ws = workbook['%s'%sheets_1[1]]
    elements = []
    rows = []
    n = 3
    while ws.cell(row=n,column=1).value is not None:
        elements.append(ws.cell(row=n,column=1).value)
        rows.append(n)
        n += 1
    print(elements, rows)
    for i in range(len(elements)):
        list = []
        m = 0
        for line in lists:
            if ('%s'%lab_number in line[0]) and ('%s'%elements[i] in line[3]) and ('(ref)' not in line[3]):
                print(line)
                list.append(line)
            m += 1
        ws.cell(row=rows[i],column=2).value = elements[i]
        if list != []:
            if str(list[0][4]) == '未校正':
                ws.cell(row=rows[i],column=3).value = '未校正'
            elif str(list[0][4]) == '####':
                ws.cell(row=rows[i],column=3).value = '超出'
            else:
                ws.cell(row=rows[i],column=3).value = float(list[0][4])*volume/quality
        else:
            ws.cell(row=rows[i],column=3).value = '未走标准曲线'
        if i == len(elements)-1:
            ws.cell(1, 3).value = lab_number
            address = os.path.abspath('.')
            workbook.save('%s\\SVHC %s.xlsx'%(address,name))

get_sample()
path=tkinter.filedialog.askopenfilenames(initialdir='Z:/Data/%s/66-01-2018-012 5110 ICP-OES/'%now_1,filetypes=[("csvfile", "*.csv")])
get_list()
action=g.ccbox('whether need to run the program again', choices=('continue','finsh'))
while  action==1:
    # get_sample()
    get_list()
    action=g.ccbox('whether need to run the program again', choices=('continue','finsh'))       
else:    
    os.startfile('.\\')

