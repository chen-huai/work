#!/usr/bin/env python
# -*- coding: utf-8 -*-


 
import tkinter as tk  # 使用Tkinter前需要先导入
import tkinter.messagebox
import pickle
import openpyxl
import easygui as g
import os.path
import win32com.client as win32com
import time
import datetime
import calendar
time_1 = int(time.strftime('%m'))
time_2=calendar.month_abbr[time_1]
address=os.path.abspath('.')

# 第1步，实例化object，建立窗口window
window = tk.Tk()
 
# 第2步，给窗口的可视化起名字
window.title('操作界面')
 
# 第3步，设定窗口的大小(长 * 宽)
window.geometry('600x400')  # 这里的乘是小x
#第一行
l_1 = tk.Label(window, text='加班申请', font=('Arial', 16,'bold'), width=40, height=2)
l_1.place(x=45, y=0) 

#第二行
var_usr_name = tk.StringVar()
var_usr_name.set('请填写名字')
entry_usr_name = tk.Entry(window, textvariable=var_usr_name, font=('Arial', 14))
entry_usr_name.place(x=30, y=55)
entry_usr_name.focus()

def select_file():
    global path
    path=tkinter.filedialog.askopenfilenames(initialdir='.\\',filetypes=[("file", "*.xls*")])   
b_1= tk.Button(window, text='选择文件', font=('Arial', 12), width=20, height=1, command=select_file)
b_1.place(x=360, y=50)

#第三行
l_2 = tk.Label(window, text='周末为工作日的，请选择：', font=('Arial', 12), width=30, height=2) 
l_2.place(x=5, y=90)
l_3 = tk.Label(window, text='工作日为假期的，请选择：', font=('Arial', 12), width=30, height=2)
l_3.place(x=325, y=90) 


#第四行
list_1=[]
list_2=[]
def print_selection_1():
    global list_1
    list_1=[]
    for m in range(31):
        if var_1[m].get()==1 :
            list_1.append(int(m)+1) 
    print(1)     
    print(list_1)  
def print_selection_2():

    global list_2
    list_2=[]
    for m in range(31):
        if var_2[m].get()==1 :
            list_2.append(int(m)+1)
    print(2) 
    print(list_2) 
var_1 = [tk.IntVar(),tk.IntVar(),tk.IntVar() ,tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar() ,tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar() ,tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar()]       
for i in range(31):
    if i < 7:
        c_i= tk.Checkbutton(window, text='%s'%(int(i)+1),variable=var_1[i], onvalue=1, offvalue=0, command=print_selection_1)
        c_i.place(x=int(i)*40, y=120)
    elif 7<=i<=13:     
        c_i= tk.Checkbutton(window, text='%s'%(int(i)+1),variable=var_1[i], onvalue=1, offvalue=0, command=print_selection_1)
        c_i.place(x=(int(i)-7)*40, y=140)
    elif 14<=i<=20:     
        c_i= tk.Checkbutton(window, text='%s'%(int(i)+1),variable=var_1[i], onvalue=1, offvalue=0, command=print_selection_1)
        c_i.place(x=(int(i)-14)*40, y=160)
    elif 21<=i<=27:
        c_i= tk.Checkbutton(window, text='%s'%(int(i)+1),variable=var_1[i], onvalue=1, offvalue=0, command=print_selection_1)
        c_i.place(x=(int(i)-21)*40, y=180)   
    else:
        c_i= tk.Checkbutton(window, text='%s'%(int(i)+1),variable=var_1[i], onvalue=1, offvalue=0, command=print_selection_1)
        c_i.place(x=(int(i)-28)*40, y=200)
var_2 = [tk.IntVar(),tk.IntVar(),tk.IntVar() ,tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar() ,tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar() ,tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar(),tk.IntVar()]      
for n in range(31):
    if n < 7:
        c_n= tk.Checkbutton(window, text='%s'%(int(n)+1),variable=var_2[n], onvalue=1, offvalue=0, command=print_selection_2)
        c_n.place(x=(int(n)+8)*40, y=120)
    elif 7<=n<=13:     
        c_n= tk.Checkbutton(window, text='%s'%(int(n)+1),variable=var_2[n], onvalue=1, offvalue=0, command=print_selection_2)
        c_n.place(x=(int(n)+1)*40, y=140)
    elif 14<=n<=20:     
        c_n= tk.Checkbutton(window, text='%s'%(int(n)+1),variable=var_2[n], onvalue=1, offvalue=0, command=print_selection_2)
        c_n.place(x=(int(n)-6)*40, y=160)
    elif 21<=n<=27:
        c_n= tk.Checkbutton(window, text='%s'%(int(n)+1),variable=var_2[n], onvalue=1, offvalue=0, command=print_selection_2)  
        c_n.place(x=(int(n)-13)*40, y=180) 
    else:
        c_n= tk.Checkbutton(window, text='%s'%(int(n)+1),variable=var_2[n], onvalue=1, offvalue=0, command=print_selection_2) 
        c_n.place(x=(int(n)-20)*40, y=200)       
 
#第五行
l_4= tk.Label(window,width=80, height=4,text='empty')         
l_4.place(x=15, y=250)
t = tk.Text(window,width=40,height=3)
t.place(x=15, y=330)

def jia_ban():
    global name
    name=var_usr_name.get()
    
    l_4.config(text='''名字：%s
周末为工作日：%s
工作日为假期：%s'''%(name,list_1,list_2))
    get_data()
    
    
b_2= tk.Button(window, text='加班确认', font=('Arial', 12), width=20, height=1, command=jia_ban)
b_2.place(x=360, y=335)

def get_data():
    excel = win32com.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.Application.DisplayAlerts = False   
    wb = openpyxl.load_workbook(os.path.join(os.getcwd(),'%s'%path))
    ws=wb.get_sheet_by_name('Sheet1')
    row_1=2
    wb_2=openpyxl.load_workbook('%s\\Overtimes Application Form.xlsx'%(address))
    ws_2=wb_2.get_sheet_by_name('加班申请表-正式&派遣&外包')
    while ws.cell(row=row_1,column=3).value!=name:
        row_1+=1
    while ws.cell(row=row_1,column=3).value==name:
        department=ws.cell(row=row_1,column=4).value
        date=ws.cell(row=row_1,column=5).value
        date_time=date.split('-')
        week=ws.cell(row=row_1,column=6).value  
        start_time=ws.cell(row=row_1,column=7).value
        end_time=ws.cell(row=row_1,column=8).value
        row_2=10
        ws_2.cell(row=5,column=7).value=name
        ws_2.cell(row=6,column=7).value=time_2
        ws_2.cell(row=6,column=3).value=department
        print(row_1,date)
        if (start_time==''):
            t.insert('end', '%s上班时间没打卡\n'%date) 
            start_time='12:30'
        elif (end_time==''): 
            t.insert('end', '%s下班时间没打卡\n'%date)
            end_time='17:30'
            print(end_time)          
        else : 
            a=datetime.datetime.strptime(str(start_time),"%H:%M")   
            b=datetime.datetime.strptime(str(end_time),"%H:%M")
            c=int((b-a).seconds)
            d=c/3600-9 
            while  ws_2.cell(row=row_2,column=1).value is not None:
                row_2+=1          
            if week=='Mon' or week=='Tue' or week=='Wed' or week=='Thu' or week=='Fri':
                if  int(date_time[2]) in list_2:
                    print(date_time,row_1,end_time,week,111)
                    ws_2.cell(row=row_2,column=1).value='Weekend'
                    ws_2.cell(row=row_2,column=2).value=date
                    ws_2.cell(row=row_2,column=3).value=start_time
                    ws_2.cell(row=row_2,column=4).value=end_time
                    ws_2.cell(row=row_2,column=5).value=(c/3600-1)
                elif d>0.91:
                    print(date_time,row_1,end_time,week,222) 
                    ws_2.cell(row=row_2,column=1).value='Weekday'
                    ws_2.cell(row=row_2,column=2).value=date
                    ws_2.cell(row=row_2,column=3).value='17:30'
                    ws_2.cell(row=row_2,column=4).value=end_time
                    ws_2.cell(row=row_2,column=5).value=d
            else:
                if  int(date_time[2]) in list_1:
                    if d>0.91:
                        print(date_time,row_1,end_time,week,333)
                        ws_2.cell(row=row_2,column=1).value='Weekday'
                        ws_2.cell(row=row_2,column=2).value=date
                        ws_2.cell(row=row_2,column=3).value='17:30'
                        ws_2.cell(row=row_2,column=4).value=end_time
                        ws_2.cell(row=row_2,column=5).value=d
                else: 
                    print(date_time,row_1,end_time,week,444)
                    ws_2.cell(row=row_2,column=1).value='Weekend'
                    ws_2.cell(row=row_2,column=2).value=date
                    ws_2.cell(row=row_2,column=3).value=start_time
                    ws_2.cell(row=row_2,column=4).value=end_time
                    ws_2.cell(row=row_2,column=5).value=(c/3600-1) 
        row_1+=1
        row_2+=1
    wb_2.save('%s\\%s Overtimes Application Form.xlsx'%(address,name))    
    tkinter.messagebox.showinfo( message='完成，请输入下一位')
window.mainloop()
