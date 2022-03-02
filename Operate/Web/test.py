import pandas as pd
from openpyxl import *

# def analysisExcel():
#     # pandas 适合分析数据
#     excelData = pd.read_excel('C:\Data\python\work\Operate\data\model\SVHC_DCU.xlsx',sheet_name=None)
#     print(list(excelData.keys()))
#     for j in excelData.keys():
#         print(j)
#     res = excelData
#
# analysisExcel()


# def readWorkbook():
#     wb =load_workbook('C:\Data\python\work\Operate\data\model\SVHC_DCU.xlsx')
#     sheetnames = wb.sheetnames
#     print(wb.sheetnames)
#     for each in sheetnames:
#         sheet = wb[each]
#         print(sheet)
#         dataRange = sheet.dimensions
#         print(dataRange)
#         column = sheet['A']
#         print(column.dimensions)
#
# readWorkbook()

class ExcelWorkbook:
    def __init__(self,data = {}):
        self.data = data
    def createWorkbook(self):
        global wb
        global ws
        wb = Workbook()
        ws = wb.active
        # 更改工作表ws的title
        ws.title = 'test_sheet1'
        # 对ws的单个单元格传入数据
        ws['A1'] = '国家'
        ws['B1'] = '首都'

    def writeSheet(self):
        data_excel = []
        # 将字典中的每对数据（键，值）以列表形式传入data_excel列表
        for each in self.data:
            data_excel.append([each, self.data[each]])
        # 将data_excel列表内的内容存入工作表
        for each in data_excel:
            ws.append(each)
        # 注意：上述两个append方法是意义完全不同的两个方法

    def saveExcel(self):
        wb.save('test.xlsx')


data = {
            '中国': '北京',
            '韩国': '首尔',
            '日本': '东京',
            '泰国': '曼谷',
            '马来西亚': '吉隆坡',
            '越南': '河内',
            '朝鲜': '平壤',
            '印度': '新德里'
        }
excelWorkbook = ExcelWorkbook(data)
excelWorkbook.createWorkbook()
excelWorkbook.writeSheet()
excelWorkbook.saveExcel()